#!/usr/bin/env python
# -*- coding: utf-8 -*-
import base64
from openpyxl import Workbook
from openpyxl import load_workbook
import pyautogui
from selenium import webdriver
import time
import os

os.chdir(os.getcwd())
wb =load_workbook('qingdan.xlsx')

# ws = wb.sheetnames
ws = wb.active
# wb = load_workbook(self.filename) #打开Excel
#         sheet = wb[self.sheetname] #定位表单
test_data= [] #创建一个空列表
for row in range(2, 242):
    # sub_data = {}
    for column in range(3, 4):
        sub_data = ws.cell(row, column).value  #Excel的第一行数据作为字典的key；
        # print(sub_data)
        test_data.append(sub_data) #将每行的数据循环加到列表中
    # print(len(test_data))
time.sleep(5)
for name in test_data:
    # 点击合并员工按钮
    pyautogui.click(780, 265)

    time.sleep(1)
    # 点击合并员工按钮
    pyautogui.click(630, 395)

    time.sleep(1)
    # 点击输入框，输入人员名字
    pyautogui.click(1350, 386)
    pyautogui.typewrite(name)
    time.sleep(1)
    #点击域用户组
    pyautogui.click(630, 395)
    time.sleep(1)
    # 选择用户
    pyautogui.click(820, 460)
    time.sleep(1)
    # 点击确认
    pyautogui.click(1300, 815)
    time.sleep(1)
    # 点击合并到的员工
    pyautogui.click(630, 630)

    time.sleep(1)
    # 点击输入框，输入人员名字
    pyautogui.click(1350, 386)
    pyautogui.typewrite(name)
    time.sleep(1)
    # 点击域用户组
    pyautogui.click(630, 420)
    time.sleep(1)
    # 选择用户
    pyautogui.click(820, 460)
    time.sleep(1)
    # 点击确认
    pyautogui.click(1300, 815)
    time.sleep(1)
    # 点击保存
    pyautogui.click(1200, 845)
    # time.sleep(5)
    while True:
        if pyautogui.pixelMatchesColor(1137, 498, (51, 203, 204), tolerance=5):
            # time.sleep(2)
            pyautogui.click(1137, 498)
            break
    pyautogui.click(1327, 307)
    time.sleep(1)
    input('输入任意键继续')



# print(test_data)
# print(ws.values)
# user = ws['A2':'A241']
# print(user)









""""""