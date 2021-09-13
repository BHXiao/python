#!/usr/bin/env python
# -*- coding: utf-8 -*-


##################################################
# '''
# 将文件的行变成列   列变成行
# '''
# #
# from openpyxl import load_workbook
#
# wb = load_workbook('qingdan.xlsx')
# ws = wb.active
# # data = []
# row = ws.max_row
# column = ws.max_column
# for x in range(1, row + 1):
#     for y in range(1, column + 1):
#         # date = ws.cell(row=x,column=y).value
#         if x > y:  # 防止x行y列这块范围的数据再次倒过来
#             ws.cell(row=x, column=y).value , ws.cell(row=y,column=x).value =\
#                     ws.cell(row=y,column=x).value, ws.cell(row=x, column=y).value
#         # ws.cell(row=y, column=x).value = date
# wb.save('qingdan1.xlsx')

#########################################################################
from openpyxl import load_workbook
# 打开窗口选择文件
import easygui

'''
创建邮件信息整理成能导入的文件
'''
open_file = easygui.fileopenbox(title='选择原始信息文件', default=r'D:\Users\5006554\Desktop\*.xlsx')
wb = load_workbook(open_file)
data = ['序号', '姓名（必填）', '帐号（必填）', '密码（必填）', '部门', '是否同时创建邮箱',
        '邮箱别名', '企业通讯录中显示的邮箱地址', '关联邮件列表地址', '重新登录修改密码设置',
        'webmail端强制开启二次登录验证', '绑定手机号码', '工号', '联系手机', '电话/分机', '传真',
        '国籍', '证件类型', '证件号码', '有效期', '备注'
        ]
# 表的列数
column = len(data)
# 表的行数
row = wb.active.max_row
# 新加一个列表。列表推导
temp_list = ['' for d in range(column) if d >= 0]
# 最后要写入文件的数据
file = ''
# 定义一个字典，为部门定值
department = {
    '行政部': '行政部',
    'IT部': 'IT部',
    '财务部': '财务部',
    '总办': '总办',
    '法务部': '法务部',
    '人力资源部': '人力资源部',
    '政府项目部': '政府项目部',
    '销售管理部': '营销系统/商务部',
    '国内用户服务部': '营销系统/用服部',
    '国内销售部': '营销系统/国内销售部',
    '国际销售部': '营销系统/International sales department',
    'IVD国际销售部': '营销系统/International sales department',
    '技术支持服务部': '营销系统/用服部',
    '用户服务部': '营销系统/用服部',
    '国际用户服务部': '营销系统/国际用户服务部',
    '国际市场部': '营销系统/市场部',
    '仪器工程部': '生产及供应链系统/仪器制造部/仪器工程部',
    '生命科学-研发一部': '生命科学/生命科学研发部;生命科学',
    '生命科学-产品管理部': '生命科学',
    '生命科学-研发三部': '生命科学/生命科学研发部;生命科学',
    '生命科学-质量法规部': '质量法规部;生命科学',
    '生命科学研发三部': '生命科学/生命科学研发部;生命科学',
    '生命科学-耗材生产部': '生命科学',
    '生命科学生产部': '生命科学',
    '生命科学质检部': '生命科学',
    '生命科学质量法规部': '质量法规部;生命科学',
    '生命科学研发一部': '生命科学/生命科学研发部;生命科学',
    '生命科学产品管理部': '生命科学',
    '采购商务部': '生产及供应链系统/采购部/采购商务部',
    '试剂生产部': '生产及供应链系统/试剂制造部/试剂生产部',
    # '临床医学部': '研发系统/临床医学部',
    # '质量法规部': '研发系统/质量法规部',
    # '中央研发部': '研发系统/中央研发部',
    # '产品安全部（中央研发）': '研发系统/中央研发部',
    # '包装设计部（中央研发）': '研发系统/中央研发部',
    # '可靠性技术部（中央研发）': '研发系统/中央研发部',
    # '工业设计部（中央研发）': '研发系统/中央研发部',
    # '研发管理部（中央研发）': '研发系统/中央研发部',
    # '整机测试部（中央研发）': '研发系统/中央研发部',
    # '移动互联部（中央研发）': '研发系统/中央研发部',
    # '算法研究部（中央研发）': '研发系统/中央研发部',
    # 'IVD应用部': '研发系统/临床医学部',
    # '临床应用部': '研发系统/临床医学部',
    # '产品管理部（IVD仪器）': '研发系统/体外诊断仪器事业部',
    # '产品系统部（IVD仪器）': '研发系统/体外诊断仪器事业部',
    # '机械工程部（IVD仪器）': '研发系统/体外诊断仪器事业部',
    # '硬件研发部（IVD仪器）': '研发系统/体外诊断仪器事业部',
    # '软件开发部（IVD仪器）': '研发系统/体外诊断仪器事业部',
    # '试剂研发一部': '研发系统/体外诊断试剂事业部',
    # '试剂研发二部': '研发系统/体外诊断试剂事业部',
    # '业务管理部（呼吸）': '研发系统/呼吸事业部',
    # '产品系统部（呼吸）': '研发系统/呼吸事业部',
    # '机械工程部（呼吸）': '研发系统/呼吸事业部',
    # '软件开发部（呼吸）': '研发系统/呼吸事业部',
    # '产品系统部（输注）': '研发系统/输注事业部',
    # '产品需求部（输注）': '研发系统/输注事业部',
    # '南京研发中心（输注）': '研发系统/输注事业部',
    '南京研发中心': '研发系统/输注事业部',
    # '机械工程部（输注）': '研发系统/输注事业部',
    # '硬件研发部（输注）': '研发系统/输注事业部',
    # '软件开发部（输注）': '研发系统/输注事业部',
    # '软件测试部（输注）': '研发系统/输注事业部',
    '输注事业部': '研发系统/输注事业部',
    '默认部门': '默认部门',
    # '': '',
    # '': '',
    # '': '',
}

for i in range(row):
    if wb.active.cell(row=i + 2, column=1).value:
        # 序号
        temp_list[0] = '\n' + str(i + 1)
        # 姓名
        temp_list[1] = wb.active.cell(row=i + 2, column=1).value
        # 账号
        temp_list[2] = wb.active.cell(row=i + 2, column=8).value
        # 密码
        temp_list[3] = 'Aa123456'
        # 定位文件中部门为空或字典中无key值的位置
        print(i + 2)
        # 部门
        temp_list[4] = department[wb.active.cell(row=i + 2, column=5).value]
        # 强制启用
        temp_list[10] = '强制开启'
        # 手机号
        temp_list[11] = wb.active.cell(row=i + 2, column=9).value
        # 工号
        temp_list[12] = wb.active.cell(row=i + 2, column=2).value
        # 最后一个回车,  csv文件写入时回车会自动下一行占用第一格
        # temp_list[-1] = '\n'
        data.extend(temp_list)
# 列表里有数字,要先转为字符
file = ','.join('%s' % id for id in data)
save_file = easygui.fileopenbox(title='保存文件为', default=r'D:\Users\5006554\Desktop\*.csv')
f = open(save_file, 'w')
f.write(file)
f.close()
##################################################################################################
# import copy
#
# '''
# 拷备的不同
# '''
# a = [1, 2, 3, 4, ['a', 'b']]
#
# b = a  # 赋值
# c = a[:]  # 浅拷贝
# d = copy.copy(a)  # 浅拷贝
# e = copy.deepcopy(a)  # 深拷贝
#
# a.append(5)
# a[4].append('c')
# a[1] = 5
#
# print('a=', a)
# print('b=', b)
# print('c=', c)
# print('d=', d)
# print('e=', e)
# # a= [1, 5, 3, 4, ['a', 'b', 'c'], 5]
# # b= [1, 5, 3, 4, ['a', 'b', 'c'], 5]
# # c= [1, 2, 3, 4, ['a', 'b', 'c']]
# # d= [1, 2, 3, 4, ['a', 'b', 'c']]
# # e= [1, 2, 3, 4, ['a', 'b']]
##################################################################################
# 99乘法表
# for i in range(1,10):
#     for j in range(1, i+1):
#         print('%d*%d=%2ld '%(j, i, i*j), end='')
#     print()
#####################################################################################
# import time
# for i in range(4):
#     print(str(int(time.time()))[-2:])
#     time.sleep(1)
####################################################################################
# thislist = list(("apple", "banana", "cherry"))  # 双括号（列表期望最多1个参数，得到3个）
# print(thislist)
############################################################################################
# def myFunc(e):
#   return e['year']
#
# cars = [
#   {'car': 'Porsche', 'year': 1963},
#   {'car': 'Audi', 'year': 2010},
#   {'car': 'BMW', 'year': 2019},
#   {'car': 'Volvo', 'year': 2013}
# ]
#
# cars.sort(key=myFunc)
# print(cars)
# # [{'car': 'Porsche', 'year': 1963}, {'car': 'Audi', 'year': 2010}, {'car': 'Volvo', 'year': 2013}, {'car': 'BMW', 'year': 2019}]

###############################################################################################
# def myFunc(_):
#   print(len(_))
#   return len(_)  # 元素中字母个数(括号中可以是任何非空字符)，用来代表元素
#
# cars = ['Porsche', 'Audi', 'BMW', 'Volvo']
#
# cars.sort(key=myFunc)
# print(cars)
# # 7
# # 4
# # 3
# # 5
# # ['BMW', 'Audi', 'Volvo', 'Porsche']
#############################################################################
# 匿名函数
# x = lambda a: a + 10
# print(x(7))
# def myfunc(n):
#   return lambda a : a * n
#
# mydoubler = myfunc(2)
#
# print(mydoubler(11))

######################################################################################
# 控制浏览调整邮箱分组
# import easygui
# import time
# from selenium import webdriver
# # from openpyxl import load_workbook
# import pandas as pd
# open_file = easygui.fileopenbox(title='选择原始信息文件', default=r'D:\Users\5006554\Desktop\*.xlsx')
#
# wb = pd.read_excel(open_file)
# ws = wb[['姓名','所属组织']]
# driver = webdriver.Chrome()
# # time.sleep(5)
# driver.get('https://qiye.163.com/login/')
# driver.find_element_by_id("switchAdminCtrl").click()
# driver.find_element_by_id("adminname").send_keys('ben.xiao@medcaptain.com')
# driver.find_element_by_id("adminpwd").send_keys('nihao,8090')
# driver.find_element_by_xpath('//*[@id="loginBlock"]/div[2]/form[2]/div[5]/button').click()
# # print(driver.window_handles)
# input()
# # driver.find_element_by_xpath('/html/body/div[5]/div/div/div/button').click()
# driver.find_element_by_xpath('//*[@id="headTabs"]/div[1]/ul/li[3]/a/tab-heading/span').click()
# time.sleep(1)
# for x,b in ws.values:
#
#     driver.find_element_by_xpath('/html/body/div[3]/div/div/div/div[2]/div'
#                                  '/div[1]/form/div[2]/div[2]/div/input').clear()
#     driver.find_element_by_xpath('/html/body/div[3]/div/div/div/div[2]/div'
#                                  '/div[1]/form/div[2]/div[2]/div/input').send_keys(x)
#     driver.find_element_by_xpath('/html/body/div[3]/div/div').click()
#     driver.find_element_by_xpath('/html/body/div[3]/div/div/div/div[2]/div'
#                                  '/div[2]/div[1]/table/tbody[2]/tr/td[1]/input').click()
#     driver.find_element_by_xpath('/html/body/div[3]/div/div/div/div[2]/div/div[1]'
#                                  '/form/div[2]/div[1]/div[2]/button').click()
#     time.sleep(0.5)
#     driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div[2]'
#                                  '/div/div[1]/div[2]/div[1]/div[1]/input').send_keys(b)
#     driver.find_element_by_xpath('//*[@id="searchUnitTreeContainer"]'
#                                  '/ul/li/label/span[2]').click()
#     driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div[2]'
#                                  '/div/button[1]').click()
#     driver.find_element_by_xpath('/html/body/div[5]/div/div/div/div[3]'
#                                  '/button[1]').click()
#     driver.find_element_by_xpath('//*[@id="searchUnitTreeContainer"]'
#                                  '/ul/li/label/span[2]').click()
#     driver.find_element_by_xpath('//*[@id="searchUnitTreeContainer"]'
#                                  '/ul/li/label/span[2]').click()
#     time.sleep(4)
# time.sleep(4)
# driver.quit()
# ######################################################################################################################
# import pyperclip 复制后 pyautogui 找不到窗口句柄问题
# import win32gui
# import win32api
# import pyautogui
#
# # from pymouse import PyMouse
# hwnd_title = {}
#
#
# def get_all_hwnd(hwnd, mouse):
#     if (win32gui.IsWindow(hwnd) and win32gui.IsWindowEnabled(hwnd) and win32gui.IsWindowVisible(hwnd)):
#         hwnd_title.update({hwnd: win32gui.GetWindowText(hwnd)})
#
#
# win32gui.EnumWindows(get_all_hwnd, 0)
#
# # m = PyMouse()
#
# for h, t in hwnd_title.items():
#     if t:
#         print(h, t)
#         if 'Mozilla Firefox' in t:
#             win32gui.GetWindow(h)
#             # left, top, right, bottom = win32gui.GetWindowRect(h)
#             # print(left, top, right, bottom)
#             # pyautogui.click(right - 206, bottom - 31)


######################################################################################################################
# 验证上面猜想(结果达不到我想要的结果)try 出错只是会不报错退出
# import win32gui
# import time
# from openpyxl import load_workbook
# import pyperclip
# import pyautogui
# import easygui
#
# # import pyHook
#
#
# # def get_all_hwnd(hwnd, mouse):
# #     if (win32gui.IsWindow(hwnd) and win32gui.IsWindowEnabled(hwnd) and win32gui.IsWindowVisible(hwnd)):
# #         hwnd_title.update({hwnd: win32gui.GetWindowText(hwnd)})
# pyautogui.FAILSAFE = True  # 设置保护措施
# pyautogui.PAUSE = 0.5      # 默认延迟是0.1 改为0.5
# hwnd_title = {}
# file = easygui.fileopenbox(title='选择原始信息文件', default=r'D:\Users\5006554\Desktop\*.xlsx')
# wb = load_workbook(file)
# ws = wb.active  # 没有括号
# # 创建一个“钩子”管理对象
# # hm = pyHook.HookManager()
# for i in range(1, ws.max_row):
#     name = ws.cell(row=i + 1, column=1).value
#     print(name)
#
#     # win32gui.EnumWindows(get_all_hwnd, 0)  # 生成一个窗口句柄的字典，存放所有打开的窗口
#     # for h, t in hwnd_title.items():
#     #     if t:
#     #         print(h, t)
#     #         if 'Mozilla Firefox' in t:
#     #             # win32gui.GetWindow(int(h, base=10), t)
#     #             # 窗体前端显示
#     #             time.sleep(20)
#     #
#     #             win32gui.SetForegroundWindow(h)
#     #             time.sleep(20)
#     #             break
#     try:
#         print(i)
#         pyautogui.click(1630, 316)
#         pyperclip.copy(name)  # 这里切换了窗口
#         print('开始全选粘贴')
#
#         pyautogui.hotkey('ctrl', 'a')  # 这里只是操作键盘，所以不报错
#         pyautogui.hotkey('ctrl', 'v')
#         print('回车')
#         pyautogui.press('enter')
#         print('第%d次最后点击'%i)
#         pyautogui.click(1000, 700)
#     except:
#         print('点击出错')
#     finally:
#         print('再次点击')
#         # pyautogui.click(1000, 700)
#######################################################################################################################
# # 迭代器
# class MyNumbers:
#   def __iter__(self):
#     self.a = 1
#     return self
#
#   def __next__(self):
#     x = self.a
#     self.a += 1
#     return x
#
# myclass = MyNumbers()
# myiter = iter(myclass)
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
# print(next(myiter))
#####################################################################################################################
# import re
#
# str = "China is a great country"
# x = re.search(r"\bC\w+", str)  # 匹配C开头的单词
# print(x)
####################################################################################################################
# import numpy as np

# print("%10.30f"%(0.1 * 3))
# print([1,2,3,4,5,6])
# arr = np.array([1,2,3,4,5,6])
# print(arr)
# print(np.__version__)
# arr = np.array(1, ndmin=5)  # 标量  就是常量  加【】就是维
# print(arr)
# print('number of dimensions :', arr.ndim)
# arr = np.array([1, 2, 3, 4], ndmin=5)
#
# print(arr)
# print('number of dimensions :', arr.ndim)
# arr = np.array([[[1,2,3],[4,5,6]],[[8,8,8],[0,2,5]]])
#
# print(arr)
# print(arr.shape)
# arr = np.array([1, 2, 3, 4], ndmin=5)
#
# print(arr)
# print('shape of array :', arr.shape)
# arr = np.array([1, 2, 3, 4, 5, 6, 7, 8])
#
# newarr = arr.reshape(2, 2,2)
#
# print(newarr)
# arr = np.array([[[1, 10, 8], [4, 7, 6]], [[7, 3, 9], [10, 7, 12]]])
#
# for x in arr:
#     print('##########')
#     print(x.base)   #  元素返回整个数组，确认是视图
#     print('##########')
# arr = np.array([[1, 2, 3, 4], [5, 6, 7, 8]])
#
# for x in np.nditer(arr[:, ::2]):  # arr[:][::2]
#   print(x)
# for idx, x in np.ndenumerate(arr):
#   print(idx, x)
# print(arr[0,0].argsort()[-1])
# print(arr[0,0,arr[0,0].argsort()[-1]])
# # print(np.random.seed(666))
# print(arr.shape)
#######################################################################################################################
# import pandas as pd
#
# se = pd.Series([1, 3, 4, 7, 9], index=['b', 'n', 'm', 'i', 'o'])
# se1 = pd.Series([9, 7, 4, 3, 1], index=['a', 'b', 'n', 'm', 'i'])
# print(se.head(2))
# print(se1.head(3))  # 默认是5
# # print((se + se1).astype(int))  # 报错是因为有空值，不支持转换成正数
# # print((se + se1).astype(int))
